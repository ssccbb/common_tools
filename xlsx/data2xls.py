import openpyxl
from openpyxl import Workbook

f = '/Users/sung/Downloads/吉州区教体系统疫情防控涉疫人员台账.xlsx'


class Excel:

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_excel(self):
        workbook = openpyxl.load_workbook(self.file_name)
        sheet = workbook[self.sheet_name]
        res = list(sheet.rows)
        cases = []
        title = [i.value for i in res[1]]
        for item in res[2:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases
        pass

    def creat_excel(self):
        wb = Workbook()
        ws = wb.active
        # 设置单元格内容
        ws['A1'] = 42
        # 设置一行内容
        ws.append([1, 2, 3])
        wb.save("sample.xlsx")
        pass


e = Excel(f, 'sheet1')
print(e.read_excel())
